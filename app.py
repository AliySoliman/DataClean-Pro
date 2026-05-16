import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
import re
import zipfile
import chardet
from datetime import datetime

st.set_page_config(
    page_title="DataClean Pro",
    page_icon="⬡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Design system ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=JetBrains+Mono:wght@300;400;600&display=swap');

:root {
    --bg:        #0d0f12;
    --surface:   #141720;
    --border:    #1e2330;
    --accent:    #00ff87;
    --accent2:   #00b8ff;
    --warn:      #ffb703;
    --danger:    #ff4757;
    --purple:    #bd93f9;
    --text:      #e2e8f0;
    --muted:     #64748b;
    --font-head: 'Syne', sans-serif;
    --font-mono: 'JetBrains Mono', monospace;
}

html, body, [class*="css"], .stApp {
    background-color: var(--bg) !important;
    color: var(--text) !important;
    font-family: var(--font-mono) !important;
}

#MainMenu, footer, header { visibility: hidden; }
.stDeployButton { display: none; }

.block-container { padding: 2rem 3rem !important; max-width: 1200px !important; }

.hero {
    display: flex; align-items: flex-end; gap: 1.5rem;
    padding: 2.5rem 0 2rem; border-bottom: 1px solid var(--border); margin-bottom: 2rem;
}
.hero-logo {
    font-family: var(--font-head); font-size: 3rem; font-weight: 800;
    color: var(--accent); letter-spacing: -2px; line-height: 1;
}
.hero-logo span { color: var(--text); }
.hero-sub { font-size: 0.72rem; color: var(--muted); letter-spacing: 0.15em; text-transform: uppercase; padding-bottom: 0.35rem; }

.pill {
    display: inline-block; font-size: 0.65rem; font-family: var(--font-mono);
    font-weight: 600; letter-spacing: 0.08em; padding: 0.2rem 0.65rem;
    border-radius: 2px; text-transform: uppercase;
}
.pill-green  { background: rgba(0,255,135,0.12); color: var(--accent);  border: 1px solid rgba(0,255,135,0.3); }
.pill-blue   { background: rgba(0,184,255,0.12); color: var(--accent2); border: 1px solid rgba(0,184,255,0.3); }
.pill-warn   { background: rgba(255,183,3,0.12); color: var(--warn);    border: 1px solid rgba(255,183,3,0.3); }
.pill-danger { background: rgba(255,71,87,0.12); color: var(--danger);  border: 1px solid rgba(255,71,87,0.3); }
.pill-purple { background: rgba(189,147,249,0.12);color: var(--purple); border: 1px solid rgba(189,147,249,0.3); }

.stat-grid { display: grid; grid-template-columns: repeat(4,1fr); gap: 1px; background: var(--border); margin: 1.25rem 0; }
.stat-card { background: var(--surface); padding: 1rem 1.25rem; display: flex; flex-direction: column; gap: 0.25rem; }
.stat-label { font-size: 0.6rem; text-transform: uppercase; letter-spacing: 0.15em; color: var(--muted); }
.stat-value { font-family: var(--font-head); font-size: 2rem; font-weight: 800; line-height: 1; }
.stat-delta { font-size: 0.7rem; }
.delta-neg { color: var(--danger); }
.delta-neu { color: var(--muted); }

.log-entry {
    display: flex; align-items: flex-start; gap: 0.85rem;
    padding: 0.65rem 0; border-bottom: 1px solid var(--border);
    font-size: 0.78rem; line-height: 1.5;
}
.log-entry:last-child { border-bottom: none; }
.log-icon { flex-shrink: 0; width: 1.5rem; text-align: center; }
.log-text { color: var(--text); }
.log-text b { color: var(--accent); }
.log-text .colname { color: var(--accent2); }

.section-head {
    font-size: 0.6rem; text-transform: uppercase; letter-spacing: 0.2em;
    color: var(--muted); padding: 1.5rem 0 0.5rem;
    border-top: 1px solid var(--border); margin-top: 1rem;
}

.file-card { border: 1px solid var(--border); background: var(--surface); padding: 1.5rem; margin-bottom: 1.5rem; }
.file-card-header {
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 1.25rem; padding-bottom: 0.85rem; border-bottom: 1px solid var(--border);
}
.file-name { font-family: var(--font-head); font-size: 1.1rem; font-weight: 700; color: var(--text); word-break: break-all; }
.arrow { color: var(--accent); margin: 0 0.5rem; }

section[data-testid="stSidebar"] { background: var(--surface) !important; border-right: 1px solid var(--border) !important; }
section[data-testid="stSidebar"] * { color: var(--text) !important; }
section[data-testid="stSidebar"] .stSlider > div > div > div 
section[data-testid="stSidebar"] hr { border-color: var(--border) !important; }

.stButton > button {
    background: var(--accent) !important; color: #000 !important;
    font-family: var(--font-mono) !important; font-weight: 600 !important;
    font-size: 0.78rem !important; letter-spacing: 0.08em !important;
    border: none !important; border-radius: 0 !important;
    padding: 0.6rem 1.4rem !important; text-transform: uppercase !important;
}
.stButton > button:hover { opacity: 0.85 !important; }

.stDownloadButton > button {
    background: transparent !important; color: var(--accent) !important;
    font-family: var(--font-mono) !important; font-size: 0.78rem !important;
    font-weight: 600 !important; border: 1px solid var(--accent) !important;
    border-radius: 0 !important; padding: 0.55rem 1.4rem !important;
    text-transform: uppercase !important; letter-spacing: 0.08em !important;
    transition: all 0.15s !important;
}
.stDownloadButton > button:hover { background: var(--accent) !important; color: #000 !important; }

.stFileUploader { border: 1px dashed var(--border) !important; background: var(--surface) !important; border-radius: 0 !important; padding: 1rem !important; }
.stExpander { border: 1px solid var(--border) !important; background: var(--surface) !important; border-radius: 0 !important; }
.stMetric { background: transparent !important; }
.stMetric label { font-size: 0.6rem !important; text-transform: uppercase !important; letter-spacing: 0.12em !important; color: var(--muted) !important; }
.stMetric [data-testid="stMetricValue"] { font-family: var(--font-head) !important; font-size: 1.8rem !important; color: var(--text) !important; }
div[data-testid="stDataFrame"] { border: 1px solid var(--border) !important; }
.stToggle label, .stRadio label, .stCheckbox label, .stSlider label, .stTextInput label, .stSelectbox label { font-size: 0.75rem !important; }
.stAlert { border-radius: 0 !important; border-left-width: 3px !important; }
.uploadlabel { font-size: 0.6rem; text-transform: uppercase; letter-spacing: 0.2em; color: var(--muted); margin-bottom: 0.5rem; display: block; }
</style>
""", unsafe_allow_html=True)

# ── Hero ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <div>
    <div class="hero-logo">Data<span>Clean</span> <span style="color:var(--accent)">Pro</span></div>
    <div class="hero-sub">CSV → Excel · Intelligent data cleaning · Audit-ready output</div>
  </div>
  <div style="margin-left:auto;display:flex;gap:0.5rem;padding-bottom:0.4rem;">
    <span class="pill pill-green">v3.0</span>
    <span class="pill pill-blue">Production</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⬡ Settings")

    st.markdown("---")
    st.markdown("**Deduplication**")
    remove_duplicates = st.toggle("Remove duplicate rows", value=True)

    st.markdown("---")
    st.markdown("**Step 1 — Column cleaning**")
    col_drop_threshold = st.slider(
        "Drop column if % missing ≥",
        min_value=0, max_value=100, value=100,
        help="100 = only fully empty columns. Lower = stricter."
    )
    st.caption("Columns are always cleaned before rows.")

    st.markdown("---")
    st.markdown("**Step 2 — Row cleaning**")
    remove_missing = st.toggle("Handle missing row values", value=True)
    row_strategy = st.radio(
        "Strategy:",
        [
            "Fill with median (recommended)",
            "Fill with mean",
            "Fill with a fixed value",
            "Drop entire row",
            "Drop only fully-empty rows",
        ],
        index=0,
        disabled=not remove_missing,
    )
    fill_value = None
    if remove_missing and row_strategy == "Fill with a fixed value":
        fill_value = st.text_input("Fixed value:", value="0")
    st.caption("Numeric cols → median/mean. Text cols → most frequent value.")

    st.markdown("---")
    st.markdown("**Excel output**")
    sheet_name    = st.text_input("Sheet name", value="Cleaned Data")
    freeze_header = st.checkbox("Freeze header row", value=True)
    auto_width    = st.checkbox("Auto column width", value=True)
    styled_header = st.checkbox("Style header row", value=True)

    st.markdown("---")
    st.markdown("**Audit**")
    download_log = st.checkbox("Include cleaning log (.txt)", value=True)


# ── Utilities ─────────────────────────────────────────────────────────────────

FAKE_NULL_PATTERNS = re.compile(
    r"^\s*$|^(nan|none|null|n/a|na|n\.a\.|missing|unknown|"
    r"undefined|nil|#n/a|#na|-|--|---)\s*$",
    re.IGNORECASE,
)

def normalize_fake_nulls(df: pd.DataFrame) -> pd.DataFrame:
    """Replace whitespace-only and common null-string placeholders with NaN."""
    obj_cols = df.select_dtypes(include="object").columns
    for col in obj_cols:
        try:
            mask = df[col].astype(str).str.match(FAKE_NULL_PATTERNS)
            df.loc[mask, col] = np.nan
        except Exception:
            pass
    return df


def coerce_mixed_numeric(df: pd.DataFrame) -> pd.DataFrame:
    """
    For columns stored as object that are mostly numeric (>70% parseable),
    coerce to float so imputation works correctly.
    Non-parseable cells become NaN.
    """
    obj_cols = df.select_dtypes(include="object").columns
    for col in obj_cols:
        try:
            coerced = pd.to_numeric(df[col], errors="coerce")
            valid_ratio = coerced.notna().sum() / max(len(df), 1)
            if valid_ratio >= 0.70:
                df[col] = coerced
        except Exception:
            pass
    return df


def replace_infinities(df: pd.DataFrame) -> pd.DataFrame:
    """Replace +inf / -inf with NaN so median/mean doesn't propagate them."""
    num_cols = df.select_dtypes(include=[np.number]).columns
    for col in num_cols:
        try:
            df[col] = df[col].replace([np.inf, -np.inf], np.nan)
        except Exception:
            pass
    return df


def deduplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename duplicate column names (col, col.1, col.2 …) to
    col, col_1, col_2 to avoid openpyxl conflicts.
    """
    seen: dict = {}
    new_cols = []
    for c in df.columns:
        c_str = str(c).strip()
        if c_str in seen:
            seen[c_str] += 1
            new_cols.append(f"{c_str}_{seen[c_str]}")
        else:
            seen[c_str] = 0
            new_cols.append(c_str)
    df.columns = new_cols
    return df


def sanitize_sheet_name(name: str) -> str:
    """Remove characters illegal in Excel sheet names and cap at 31 chars."""
    cleaned = re.sub(r"[\\/*?\[\]:]", "", name).strip() or "Cleaned Data"
    return cleaned[:31]


def detect_encoding(raw_bytes: bytes) -> str:
    try:
        result = chardet.detect(raw_bytes)
        return result.get("encoding") or "utf-8"
    except Exception:
        return "utf-8"


def detect_delimiter(sample: str) -> str:
    counts = {d: sample.count(d) for d in [",", ";", "\t", "|"]}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


def quality_score(df_raw: pd.DataFrame, dropped_cols: list) -> int:
    if df_raw.empty:
        return 0
    total_cells = df_raw.shape[0] * df_raw.shape[1]
    missing     = df_raw.isna().sum().sum()
    dup_rows    = df_raw.duplicated().sum()
    empty_cols  = len(dropped_cols)
    penalty  = (missing    / max(total_cells, 1)) * 40
    penalty += (dup_rows   / max(len(df_raw), 1)) * 30
    penalty += (empty_cols / max(df_raw.shape[1], 1)) * 30
    return max(0, min(100, round(100 - penalty)))


def quality_color(score: int) -> str:
    return "#00ff87" if score >= 85 else "#ffb703" if score >= 60 else "#ff4757"


def quality_label(score: int) -> str:
    return "Excellent" if score >= 85 else "Fair" if score >= 60 else "Poor"


def col_type_summary(df: pd.DataFrame) -> dict:
    s = {"Numeric": 0, "Categorical": 0, "Datetime": 0}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            s["Numeric"] += 1
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            s["Datetime"] += 1
        else:
            s["Categorical"] += 1
    return {k: v for k, v in s.items() if v > 0}


def missing_heatmap(df_raw: pd.DataFrame):
    try:
        # Cap columns to avoid performance issues on very wide files
        df_sample = df_raw.iloc[:, :50] if df_raw.shape[1] > 50 else df_raw
        miss_pct = (df_sample.isna().sum() / max(len(df_sample), 1) * 100).sort_values(ascending=True)
        miss_pct = miss_pct[miss_pct > 0]
        if miss_pct.empty:
            return None
        colors = ["#00ff87" if v < 20 else "#ffb703" if v < 60 else "#ff4757" for v in miss_pct.values]
        fig = go.Figure(go.Bar(
            x=miss_pct.values, y=miss_pct.index, orientation="h",
            marker_color=colors,
            text=[f"{v:.1f}%" for v in miss_pct.values],
            textposition="outside",
            textfont=dict(family="JetBrains Mono", size=10, color="#e2e8f0"),
        ))
        fig.update_layout(
            paper_bgcolor="#141720", plot_bgcolor="#141720",
            font=dict(family="JetBrains Mono", color="#e2e8f0", size=10),
            margin=dict(l=10, r=60, t=10, b=10),
            height=max(120, len(miss_pct) * 32),
            xaxis=dict(range=[0, 115], ticksuffix="%", gridcolor="#1e2330", showline=False, tickfont=dict(size=9)),
            yaxis=dict(gridcolor="#1e2330", showline=False),
            bargap=0.35,
        )
        return fig
    except Exception:
        return None


def build_audit_log(filename: str, report: dict, score: int) -> str:
    lines = [
        "=" * 60,
        "  DataClean Pro — Cleaning Report",
        f"  File    : {filename}",
        f"  Date    : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60, "",
        f"  Quality Score (before) : {score}/100 — {quality_label(score)}", "",
        "  SUMMARY",
        f"  Encoding    : {report.get('encoding', 'utf-8')}",
        f"  Delimiter   : {repr(report.get('delimiter', ','))}",
        f"  Rows before : {report['rows_before']}",
        f"  Rows after  : {report['rows_after']}",
        f"  Cols before : {report['cols_before']}",
        f"  Cols after  : {report['cols_after']}", "",
    ]
    if report["dropped_cols"]:
        lines.append("  DROPPED COLUMNS")
        for col, pct in report["dropped_cols"]:
            lines.append(f"  - '{col}' ({pct}% missing) removed")
        lines.append("")
    if report["coerced_cols"]:
        lines.append("  TYPE-COERCED COLUMNS (mixed text+number → numeric)")
        for col in report["coerced_cols"]:
            lines.append(f"  - '{col}'")
        lines.append("")
    if report["inf_cols"]:
        lines.append("  INFINITE VALUES REPLACED WITH NaN")
        for col in report["inf_cols"]:
            lines.append(f"  - '{col}'")
        lines.append("")
    if report["removed_dup"]:
        lines += [f"  DUPLICATES : {report['removed_dup']} row(s) removed", ""]
    if report["removed_rows"]:
        lines += [f"  MISSING ROWS : {report['removed_rows']} row(s) dropped", ""]
    if report["imputed_cols"]:
        lines.append("  IMPUTED COLUMNS")
        for col, method, val in report["imputed_cols"]:
            lines.append(f"  - '{col}' filled with {method} = {val}")
        lines.append("")
    if report["col_types"]:
        lines.append("  COLUMN TYPES (after cleaning)")
        for t, n in report["col_types"].items():
            lines.append(f"  - {t}: {n}")
        lines.append("")
    lines += ["=" * 60, "  End of report", "=" * 60]
    return "\n".join(lines)


# ── Core cleaning logic ───────────────────────────────────────────────────────

def clean_dataframe(df_raw: pd.DataFrame, encoding: str = "utf-8", delimiter: str = ","):
    report = {
        "rows_before":  len(df_raw),
        "cols_before":  len(df_raw.columns),
        "dropped_cols": [],
        "coerced_cols": [],
        "inf_cols":     [],
        "removed_dup":  0,
        "removed_rows": 0,
        "imputed_cols": [],
        "rows_after":   0,
        "cols_after":   0,
        "col_types":    {},
        "encoding":     encoding,
        "delimiter":    delimiter,
    }

    df = df_raw.copy()

    # ── Pre-processing ────────────────────────────────────────────────────────
    # Strip whitespace from column names
    df.columns = [str(c).strip() for c in df.columns]
    df = deduplicate_columns(df)

    # Normalize fake null strings ("N/A", "null", "  ", etc.) → NaN
    df = normalize_fake_nulls(df)

    # Coerce mixed-type columns that are mostly numeric
    before_cols = set(df.columns)
    df_coerced  = coerce_mixed_numeric(df.copy())
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df_coerced[col]) and not pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df_coerced[col]
            report["coerced_cols"].append(col)

    # Replace infinities
    for col in df.select_dtypes(include=[np.number]).columns:
        inf_count = np.isinf(df[col]).sum()
        if inf_count > 0:
            df[col] = df[col].replace([np.inf, -np.inf], np.nan)
            report["inf_cols"].append(col)

    # ── Step 1: drop columns above missing threshold ──────────────────────────
    n_rows = len(df)
    if n_rows > 0:
        for col in list(df.columns):
            pct = df[col].isna().sum() / n_rows * 100
            if pct >= col_drop_threshold:
                report["dropped_cols"].append((col, round(pct, 1)))
                df = df.drop(columns=[col])

    if df.columns.empty:
        raise ValueError("All columns were dropped. Raise the column threshold in the sidebar.")

    # ── Step 2: duplicates ────────────────────────────────────────────────────
    if remove_duplicates:
        before = len(df)
        df = df.drop_duplicates()
        report["removed_dup"] = before - len(df)

    # ── Step 3: row missing values ────────────────────────────────────────────
    if remove_missing and len(df) > 0:

        if row_strategy == "Fill with median (recommended)":
            for col in df.columns:
                if df[col].isna().any():
                    if pd.api.types.is_numeric_dtype(df[col]):
                        # Safe median: skip if all-NaN or single value
                        non_null = df[col].dropna()
                        if len(non_null) == 0:
                            val = 0.0
                        elif len(non_null) == 1:
                            val = float(non_null.iloc[0])
                        else:
                            val = float(non_null.median())
                        df[col] = df[col].fillna(val)
                        report["imputed_cols"].append((col, "median", round(val, 4)))
                    else:
                        mode_vals = df[col].mode()
                        if not mode_vals.empty:
                            df[col] = df[col].fillna(mode_vals[0])
                            report["imputed_cols"].append((col, "mode", str(mode_vals[0])))
                        else:
                            df[col] = df[col].fillna("Unknown")
                            report["imputed_cols"].append((col, "fallback", "Unknown"))

        elif row_strategy == "Fill with mean":
            for col in df.columns:
                if df[col].isna().any():
                    if pd.api.types.is_numeric_dtype(df[col]):
                        non_null = df[col].dropna()
                        val = float(non_null.mean()) if len(non_null) > 0 else 0.0
                        df[col] = df[col].fillna(val)
                        report["imputed_cols"].append((col, "mean", round(val, 4)))
                    else:
                        mode_vals = df[col].mode()
                        if not mode_vals.empty:
                            df[col] = df[col].fillna(mode_vals[0])
                            report["imputed_cols"].append((col, "mode", str(mode_vals[0])))
                        else:
                            df[col] = df[col].fillna("Unknown")
                            report["imputed_cols"].append((col, "fallback", "Unknown"))

        elif row_strategy == "Fill with a fixed value":
            fv = fill_value if fill_value is not None else "0"
            for col in df.columns:
                if df[col].isna().any():
                    df[col] = df[col].fillna(fv)
                    report["imputed_cols"].append((col, "fixed", fv))

        elif row_strategy == "Drop entire row":
            before = len(df)
            df = df.dropna()
            report["removed_rows"] = before - len(df)

        elif row_strategy == "Drop only fully-empty rows":
            before = len(df)
            df = df.dropna(how="all")
            report["removed_rows"] = before - len(df)

    report["rows_after"] = len(df)
    report["cols_after"] = len(df.columns)
    report["col_types"]  = col_type_summary(df)
    return df, report


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_excel(df: pd.DataFrame) -> io.BytesIO:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    safe_sheet = sanitize_sheet_name(sheet_name or "Cleaned Data")
    buf = io.BytesIO()
    try:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=safe_sheet)
            ws = writer.sheets[safe_sheet]

            if styled_header:
                hdr_fill = PatternFill("solid", fgColor="00FF87")
                hdr_font = Font(name="Calibri", bold=True, color="000000", size=11)
                thin     = Side(style="thin", color="00CC6A")
                for cell in ws[1]:
                    try:
                        cell.fill      = hdr_fill
                        cell.font      = hdr_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border    = Border(bottom=thin)
                    except Exception:
                        pass

            if freeze_header:
                ws.freeze_panes = "A2"

            if auto_width:
                for col_cells in ws.columns:
                    try:
                        max_len = max(
                            (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
                            default=0,
                        )
                        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)
                    except Exception:
                        pass
    except Exception as e:
        raise RuntimeError(f"Excel write failed: {e}") from e

    buf.seek(0)
    return buf


# ── Upload ────────────────────────────────────────────────────────────────────
st.markdown('<span class="uploadlabel">Input files</span>', unsafe_allow_html=True)
uploaded_files = st.file_uploader(
    "Drop CSV files here — multiple files supported",
    type=["csv"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

if not uploaded_files:
    st.markdown("""
    <div style="padding:3rem 0;text-align:center;color:#64748b;font-size:0.78rem;letter-spacing:0.1em;text-transform:uppercase;">
        ↑ &nbsp; Upload one or more CSV files to begin
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ── Process files ─────────────────────────────────────────────────────────────
results = []

for f in uploaded_files:
    try:
        raw_bytes = f.read()
        if not raw_bytes:
            st.warning(f"**{f.name}** is empty — skipped.")
            continue

        # Encoding & delimiter detection
        enc = detect_encoding(raw_bytes)
        try:
            text_sample = raw_bytes[:8192].decode(enc, errors="replace")
        except Exception:
            text_sample = raw_bytes[:8192].decode("utf-8", errors="replace")
            enc = "utf-8"
        delim = detect_delimiter(text_sample)

        # Parse CSV
        try:
            df_raw = pd.read_csv(
                io.BytesIO(raw_bytes),
                encoding=enc,
                sep=delim,
                on_bad_lines="skip",
                low_memory=False,
            )
        except pd.errors.EmptyDataError:
            st.warning(f"**{f.name}** is empty — skipped.")
            continue
        except pd.errors.ParserError as e:
            # Fallback: try comma regardless
            try:
                df_raw = pd.read_csv(io.BytesIO(raw_bytes), encoding=enc, sep=",", on_bad_lines="skip")
                delim = ","
            except Exception:
                st.error(f"**{f.name}** could not be parsed: {e}")
                continue
        except UnicodeDecodeError:
            try:
                df_raw = pd.read_csv(io.BytesIO(raw_bytes), encoding="latin-1", sep=delim, on_bad_lines="skip")
                enc = "latin-1"
            except Exception as e2:
                st.error(f"**{f.name}** encoding error — could not read: {e2}")
                continue
        except Exception as e:
            st.error(f"**{f.name}** could not be read: {e}")
            continue

        if df_raw.empty or len(df_raw.columns) == 0:
            st.warning(f"**{f.name}** has no usable data — skipped.")
            continue

        # Compute quality score & heatmap on raw data
        preview_dropped = [
            col for col in df_raw.columns
            if df_raw[col].isna().sum() / max(len(df_raw), 1) * 100 >= col_drop_threshold
        ]
        score       = quality_score(df_raw, preview_dropped)
        heatmap_fig = missing_heatmap(df_raw)

        # Clean
        try:
            df, report = clean_dataframe(df_raw, encoding=enc, delimiter=delim)
        except ValueError as e:
            st.error(f"**{f.name}**: {e}")
            continue
        except Exception as e:
            st.error(f"**{f.name}** — unexpected cleaning error: {e}")
            continue

        if df.empty:
            st.warning(
                f"**{f.name}** — no rows remain after cleaning. "
                "Switch to a fill strategy instead of dropping rows."
            )
            continue

        # Write Excel
        try:
            excel_buf = write_excel(df)
        except Exception as e:
            st.error(f"**{f.name}** — failed to write Excel: {e}")
            continue

        # Safe output filename
        safe_stem = re.sub(r"[^\w\-]", "_", f.name.replace(".csv", ""))
        out_name  = f"{safe_stem}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        log_text  = build_audit_log(f.name, report, score)

        results.append({
            "name":        f.name,
            "out_name":    out_name,
            "report":      report,
            "score":       score,
            "heatmap":     heatmap_fig,
            "excel_bytes": excel_buf.read(),
            "log_text":    log_text,
            "df":          df,
            "df_raw":      df_raw,
        })

    except Exception as e:
        st.error(f"Unexpected error with **{f.name}**: {e}")


if not results:
    st.stop()

# ── Global summary bar ────────────────────────────────────────────────────────
total_rows_in  = sum(r["report"]["rows_before"] for r in results)
total_rows_out = sum(r["report"]["rows_after"]  for r in results)
total_cols_in  = sum(r["report"]["cols_before"] for r in results)
total_cols_out = sum(r["report"]["cols_after"]  for r in results)
total_dups     = sum(r["report"]["removed_dup"] for r in results)
avg_score      = round(sum(r["score"] for r in results) / len(results))

st.markdown(f"""
<div style="display:grid;grid-template-columns:repeat(6,1fr);gap:1px;background:var(--border);margin:1.5rem 0 2rem;">
  <div style="background:var(--surface);padding:1rem 1.25rem;">
    <div class="stat-label">Files</div>
    <div class="stat-value" style="color:var(--accent)">{len(results)}</div>
  </div>
  <div style="background:var(--surface);padding:1rem 1.25rem;">
    <div class="stat-label">Rows in</div>
    <div class="stat-value">{total_rows_in:,}</div>
  </div>
  <div style="background:var(--surface);padding:1rem 1.25rem;">
    <div class="stat-label">Rows out</div>
    <div class="stat-value" style="color:var(--accent)">{total_rows_out:,}</div>
  </div>
  <div style="background:var(--surface);padding:1rem 1.25rem;">
    <div class="stat-label">Cols dropped</div>
    <div class="stat-value" style="color:{'var(--danger)' if (total_cols_in-total_cols_out)>0 else 'var(--muted)'}">{total_cols_in-total_cols_out}</div>
  </div>
  <div style="background:var(--surface);padding:1rem 1.25rem;">
    <div class="stat-label">Duplicates</div>
    <div class="stat-value" style="color:{'var(--warn)' if total_dups>0 else 'var(--muted)'}">{total_dups:,}</div>
  </div>
  <div style="background:var(--surface);padding:1rem 1.25rem;">
    <div class="stat-label">Avg quality</div>
    <div class="stat-value" style="color:{quality_color(avg_score)}">{avg_score}</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Per-file results ──────────────────────────────────────────────────────────
# ── Per-file results ──────────────────────────────────────────────────────────
for r in results:
    rep   = r["report"]
    score = r["score"]
    qc    = quality_color(score)
    ql    = quality_label(score)

    # Build pills and deltas as plain strings to avoid f-string nesting issues
    pills = (
        f'<span class="pill pill-blue">enc: {rep.get("encoding","utf-8")}</span>'
        f'<span class="pill pill-blue">sep: {repr(rep.get("delimiter",","))}</span>'
        + "".join(f'<span class="pill pill-green">{t}: {n}</span>' for t, n in rep["col_types"].items())
        + (f'<span class="pill pill-warn">coerced: {len(rep["coerced_cols"])}</span>' if rep["coerced_cols"] else "")
        + (f'<span class="pill pill-warn">inf fixed: {len(rep["inf_cols"])}</span>' if rep["inf_cols"] else "")
    )

    r_delta   = rep["rows_before"] - rep["rows_after"]
    c_delta   = rep["cols_before"] - rep["cols_after"]
    r_d_html  = f'<div class="stat-delta delta-neg">\u2212{r_delta}</div>' if r_delta > 0 else '<div class="stat-delta delta-neu">\u2014</div>'
    c_d_html  = f'<div class="stat-delta delta-neg">\u2212{c_delta}</div>' if c_delta > 0 else '<div class="stat-delta delta-neu">\u2014</div>'

    card_html = (
        '<div class="file-card">'
        '<div class="file-card-header">'
        '<div style="flex:1;min-width:0;">'
        f'<div class="file-name">{r["name"]}<span class="arrow">\u2192</span>{r["out_name"]}</div>'
        f'<div style="margin-top:0.4rem;display:flex;gap:0.4rem;flex-wrap:wrap;">{pills}</div>'
        '</div>'
        '<div style="text-align:right;flex-shrink:0;padding-left:1rem;">'
        '<div style="font-size:0.6rem;text-transform:uppercase;letter-spacing:0.15em;color:var(--muted);">Quality score</div>'
        f'<div style="font-family:var(--font-head);font-size:2.5rem;font-weight:800;color:{qc};line-height:1.1;">{score}<span style="font-size:1rem;color:var(--muted);">/100</span></div>'
        f'<div style="font-size:0.65rem;color:{qc};">{ql}</div>'
        '</div>'
        '</div>'
        '<div class="stat-grid">'
        '<div class="stat-card">'
        '<div class="stat-label">Rows before</div>'
        f'<div class="stat-value">{rep["rows_before"]:,}</div>'
        '</div>'
        '<div class="stat-card">'
        '<div class="stat-label">Rows after</div>'
        f'<div class="stat-value" style="color:var(--accent)">{rep["rows_after"]:,}</div>'
        + r_d_html +
        '</div>'
        '<div class="stat-card">'
        '<div class="stat-label">Cols before</div>'
        f'<div class="stat-value">{rep["cols_before"]}</div>'
        '</div>'
        '<div class="stat-card">'
        '<div class="stat-label">Cols after</div>'
        f'<div class="stat-value" style="color:var(--accent)">{rep["cols_after"]}</div>'
        + c_d_html +
        '</div>'
        '</div>'
        '</div>'
    )
    st.markdown(card_html, unsafe_allow_html=True)


    # Missing value heatmap
    if r["heatmap"] is not None:
        with st.expander("Missing value map (before cleaning)", expanded=False):
            st.plotly_chart(r["heatmap"], use_container_width=True, config={"displayModeBar": False})

    # Cleaning log
    log_entries = []
    for col_name, pct in rep["dropped_cols"]:
        label = "fully empty" if pct == 100.0 else f"{pct}% missing"
        log_entries.append(("🗑️", f'Column <span class="colname">"{col_name}"</span> removed — <b>{label}</b>. No usable data in this feature.'))
    for col_name in rep["coerced_cols"]:
        log_entries.append(("🔀", f'Column <span class="colname">"{col_name}"</span> — mixed text+numbers detected, coerced to numeric.'))
    for col_name in rep["inf_cols"]:
        log_entries.append(("♾️", f'Column <span class="colname">"{col_name}"</span> — infinite values replaced with NaN before imputation.'))
    if rep["removed_dup"]:
        log_entries.append(("🔁", f'<b>{rep["removed_dup"]}</b> duplicate row(s) removed.'))
    if rep["removed_rows"]:
        log_entries.append(("🕳️", f'<b>{rep["removed_rows"]}</b> row(s) dropped due to missing values.'))
    for col_name, method, val in rep["imputed_cols"]:
        log_entries.append(("🔧", f'Column <span class="colname">"{col_name}"</span> — empty cells filled with <b>{method} = {val}</b>.'))
    if not log_entries:
        log_entries.append(("✅", "Data was already clean — no changes made."))

    entries_html = "".join(
        f'<div class="log-entry"><div class="log-icon">{icon}</div><div class="log-text">{text}</div></div>'
        for icon, text in log_entries
    )
    st.markdown(f"""
    <div class="section-head">Cleaning log</div>
    <div style="background:var(--surface);border:1px solid var(--border);padding:0.5rem 1rem;">
      {entries_html}
    </div>
    """, unsafe_allow_html=True)

    # Download row
    st.markdown('<div class="section-head">Download</div>', unsafe_allow_html=True)
    dl1, dl2, _ = st.columns([2, 2, 3])
    with dl1:
        st.download_button(
            label="⬇ Download Excel",
            data=r["excel_bytes"],
            file_name=r["out_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_{r['name']}",
            use_container_width=True,
        )
    with dl2:
        if download_log:
            st.download_button(
                label="⬇ Cleaning Log",
                data=r["log_text"].encode("utf-8"),
                file_name=r["out_name"].replace(".xlsx", "_log.txt"),
                mime="text/plain",
                key=f"log_{r['name']}",
                use_container_width=True,
            )

    # Preview with before/after null counts
    if st.toggle("Preview cleaned data", key=f"prev_{r['name']}"):
        tab1, tab2 = st.tabs(["Cleaned data", "Null count comparison"])
        with tab1:
            st.dataframe(r["df"].head(100), use_container_width=True, height=320)
        with tab2:
            null_before = r["df_raw"].isna().sum().rename("Before")
            null_after  = r["df"].reindex(columns=r["df_raw"].columns).isna().sum().rename("After")
            null_cmp    = pd.concat([null_before, null_after], axis=1).fillna(0).astype(int)
            null_cmp    = null_cmp[null_cmp["Before"] > 0]
            if null_cmp.empty:
                st.info("No missing values found in the original file.")
            else:
                st.dataframe(null_cmp, use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)


# ── Bulk ZIP ──────────────────────────────────────────────────────────────────
if len(results) > 1:
    st.markdown('<div class="section-head">Bulk export</div>', unsafe_allow_html=True)
    try:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for r in results:
                zf.writestr(r["out_name"], r["excel_bytes"])
                if download_log:
                    zf.writestr(r["out_name"].replace(".xlsx", "_log.txt"), r["log_text"].encode("utf-8"))
        zip_buf.seek(0)
        st.download_button(
            label="⬇ Download ALL as ZIP",
            data=zip_buf.read(),
            file_name=f"dataclean_export_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
            mime="application/zip",
        )
    except Exception as e:
        st.error(f"Could not create ZIP: {e}")

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style="margin-top:4rem;padding-top:1.5rem;border-top:1px solid var(--border);
     display:flex;justify-content:space-between;align-items:center;
     font-size:0.65rem;color:var(--muted);letter-spacing:0.1em;text-transform:uppercase;">
  <div>DataClean Pro &nbsp;·&nbsp; Built with Streamlit</div>
  <div>{datetime.now().strftime('%Y-%m-%d')}</div>
</div>
""", unsafe_allow_html=True)