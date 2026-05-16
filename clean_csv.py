"""
clean_csv.py — DataClean Pro (Standalone Script)
-------------------------------------------------
Converts CSV files to cleaned Excel files from the command line.
Default strategy: fill missing values with MEDIAN (numeric) or MODE (text).

Usage:
    python clean_csv.py                        # cleans all CSVs in current folder
    python clean_csv.py data.csv               # single file
    python clean_csv.py file1.csv file2.csv    # multiple files
    python clean_csv.py --input ./my_folder    # all CSVs in a folder
    python clean_csv.py --strategy drop_row    # override strategy

Strategies:
    fill_median   fill numeric with median, text with mode (DEFAULT)
    fill_mean     fill numeric with mean, text with mode
    drop_row      remove any row with a missing value
    drop_empty    remove only fully-blank rows
    fill_zero     fill everything with 0

Other options:
    --output PATH         output folder (default: ./output)
    --col-threshold INT   drop column if % missing >= this (default: 100)
    --keep-duplicates     keep duplicate rows
    --no-log              skip audit log .txt
    --sheet NAME          Excel sheet name (default: Cleaned Data)
    --fixed-value TEXT    used with --strategy fill_fixed
"""

import sys
import os
import io
import re
import argparse
import chardet
import numpy as np
import pandas as pd
from datetime import datetime
from pathlib import Path

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not found. Run: pip install openpyxl")


# ── Null-string patterns ──────────────────────────────────────────────────────
FAKE_NULL_RE = re.compile(
    r"^\s*$|^(nan|none|null|n/a|na|n\.a\.|missing|unknown|"
    r"undefined|nil|#n/a|#na|-|--|---)\s*$",
    re.IGNORECASE,
)


# ── Helpers ───────────────────────────────────────────────────────────────────
def detect_encoding(raw_bytes: bytes) -> str:
    try:
        result = chardet.detect(raw_bytes)
        return result.get("encoding") or "utf-8"
    except Exception:
        return "utf-8"


def detect_delimiter(sample: str) -> str:
    counts = {d: sample.count(d) for d in [",", ";", "\t", "|"]}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


def sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?\[\]:]", "", name).strip() or "Cleaned Data"
    return cleaned[:31]


def sanitize_filename(name: str) -> str:
    return re.sub(r"[^\w\-]", "_", name)


def normalize_fake_nulls(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.select_dtypes(include="object").columns:
        try:
            mask = df[col].astype(str).str.match(FAKE_NULL_RE)
            df.loc[mask, col] = np.nan
        except Exception:
            pass
    return df


def coerce_mixed_numeric(df: pd.DataFrame) -> tuple:
    coerced = []
    for col in df.select_dtypes(include="object").columns:
        try:
            c = pd.to_numeric(df[col], errors="coerce")
            if c.notna().sum() / max(len(df), 1) >= 0.70:
                df[col] = c
                coerced.append(col)
        except Exception:
            pass
    return df, coerced


def fix_infinities(df: pd.DataFrame) -> tuple:
    inf_cols = []
    for col in df.select_dtypes(include=[np.number]).columns:
        try:
            if np.isinf(df[col]).any():
                df[col] = df[col].replace([np.inf, -np.inf], np.nan)
                inf_cols.append(col)
        except Exception:
            pass
    return df, inf_cols


def deduplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    seen: dict = {}
    new_cols = []
    for c in df.columns:
        s = str(c).strip()
        if s in seen:
            seen[s] += 1
            new_cols.append(f"{s}_{seen[s]}")
        else:
            seen[s] = 0
            new_cols.append(s)
    df.columns = new_cols
    return df


def col_type_summary(df: pd.DataFrame) -> dict:
    s = {"Numeric": 0, "Categorical": 0, "Datetime": 0}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):      s["Numeric"] += 1
        elif pd.api.types.is_datetime64_any_dtype(df[col]): s["Datetime"] += 1
        else:                                            s["Categorical"] += 1
    return {k: v for k, v in s.items() if v > 0}


def quality_score(df: pd.DataFrame, dropped_cols: list) -> int:
    if df.empty:
        return 0
    total = df.shape[0] * df.shape[1]
    miss  = df.isna().sum().sum()
    dups  = df.duplicated().sum()
    ec    = len(dropped_cols)
    p  = (miss / max(total, 1)) * 40
    p += (dups  / max(len(df), 1)) * 30
    p += (ec    / max(df.shape[1], 1)) * 30
    return max(0, min(100, round(100 - p)))


# ── Core cleaning ─────────────────────────────────────────────────────────────
def clean_dataframe(df_raw: pd.DataFrame, args) -> tuple:
    report = {
        "rows_before":  len(df_raw),
        "cols_before":  len(df_raw.columns),
        "dropped_cols": [],
        "coerced_cols": [],
        "inf_cols":     [],
        "removed_dup":  0,
        "removed_rows": 0,
        "imputed_cols": [],
        "rows_after":   0,
        "cols_after":   0,
    }

    df = df_raw.copy()

    # Pre-processing
    df.columns = [str(c).strip() for c in df.columns]
    df = deduplicate_columns(df)
    df = normalize_fake_nulls(df)
    df, report["coerced_cols"] = coerce_mixed_numeric(df)
    df, report["inf_cols"]     = fix_infinities(df)

    # Step 1 — drop columns
    n = len(df)
    if n > 0:
        for col in list(df.columns):
            pct = df[col].isna().sum() / n * 100
            if pct >= args.col_threshold:
                report["dropped_cols"].append((col, round(pct, 1)))
                df = df.drop(columns=[col])

    if df.columns.empty:
        raise ValueError("All columns dropped. Raise --col-threshold.")

    # Step 2 — duplicates
    if not args.keep_duplicates:
        before = len(df)
        df = df.drop_duplicates()
        report["removed_dup"] = before - len(df)

    # Step 3 — rows
    if len(df) == 0:
        report["rows_after"] = 0
        report["cols_after"] = len(df.columns)
        return df, report

    strat = args.strategy

    if strat in ("fill_median", "fill_mean"):
        for col in df.columns:
            if df[col].isna().any():
                if pd.api.types.is_numeric_dtype(df[col]):
                    non_null = df[col].dropna()
                    if len(non_null) == 0:
                        val = 0.0
                    elif len(non_null) == 1:
                        val = float(non_null.iloc[0])
                    else:
                        val = float(non_null.median()) if strat == "fill_median" else float(non_null.mean())
                    df[col] = df[col].fillna(val)
                    report["imputed_cols"].append((col, strat.replace("fill_", ""), round(val, 4)))
                else:
                    mode_v = df[col].mode()
                    fv = mode_v[0] if not mode_v.empty else "Unknown"
                    df[col] = df[col].fillna(fv)
                    report["imputed_cols"].append((col, "mode", str(fv)))

    elif strat == "fill_zero":
        for col in df.columns:
            if df[col].isna().any():
                df[col] = df[col].fillna(0)
                report["imputed_cols"].append((col, "fixed", 0))

    elif strat == "fill_fixed":
        fv = getattr(args, "fixed_value", "0") or "0"
        for col in df.columns:
            if df[col].isna().any():
                df[col] = df[col].fillna(fv)
                report["imputed_cols"].append((col, "fixed", fv))

    elif strat == "drop_row":
        before = len(df)
        df = df.dropna()
        report["removed_rows"] = before - len(df)

    elif strat == "drop_empty":
        before = len(df)
        df = df.dropna(how="all")
        report["removed_rows"] = before - len(df)

    report["rows_after"] = len(df)
    report["cols_after"] = len(df.columns)
    return df, report


# ── Excel writer ──────────────────────────────────────────────────────────────
def write_excel(df: pd.DataFrame, sheet: str, output_path: Path) -> None:
    safe_sheet = sanitize_sheet_name(sheet)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=safe_sheet)
        ws = writer.sheets[safe_sheet]

        if HAS_OPENPYXL:
            hdr_fill = PatternFill("solid", fgColor="00FF87")
            hdr_font = Font(name="Calibri", bold=True, color="000000", size=11)
            thin     = Side(style="thin", color="00CC6A")
            for cell in ws[1]:
                try:
                    cell.fill      = hdr_fill
                    cell.font      = hdr_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = Border(bottom=thin)
                except Exception:
                    pass

        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            try:
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
                    default=0,
                )
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)
            except Exception:
                pass

    output_path.write_bytes(buf.getvalue())


# ── Audit log ─────────────────────────────────────────────────────────────────
def write_log(filename: str, report: dict, score: int, enc: str, delim: str, path: Path) -> None:
    lines = [
        "=" * 60,
        "  DataClean Pro — Cleaning Report",
        f"  File      : {filename}",
        f"  Date      : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60, "",
        f"  Quality Score (before) : {score}/100", "",
        "  SUMMARY",
        f"  Encoding    : {enc}",
        f"  Delimiter   : {repr(delim)}",
        f"  Rows before : {report['rows_before']}",
        f"  Rows after  : {report['rows_after']}",
        f"  Cols before : {report['cols_before']}",
        f"  Cols after  : {report['cols_after']}", "",
    ]
    if report["dropped_cols"]:
        lines.append("  DROPPED COLUMNS")
        for col, pct in report["dropped_cols"]:
            lines.append(f"  - '{col}' ({pct}% missing)")
        lines.append("")
    if report["coerced_cols"]:
        lines.append("  TYPE-COERCED TO NUMERIC")
        for col in report["coerced_cols"]:
            lines.append(f"  - '{col}'")
        lines.append("")
    if report["inf_cols"]:
        lines.append("  INFINITE VALUES REPLACED")
        for col in report["inf_cols"]:
            lines.append(f"  - '{col}'")
        lines.append("")
    if report["removed_dup"]:
        lines += [f"  DUPLICATES : {report['removed_dup']} removed", ""]
    if report["removed_rows"]:
        lines += [f"  ROWS DROPPED : {report['removed_rows']}", ""]
    if report["imputed_cols"]:
        lines.append("  IMPUTED")
        for col, method, val in report["imputed_cols"]:
            lines.append(f"  - '{col}' → {method} = {val}")
        lines.append("")
    lines += ["=" * 60, "  End of report", "=" * 60]
    path.write_text("\n".join(lines), encoding="utf-8")


# ── Process one file ──────────────────────────────────────────────────────────
def process_file(csv_path: Path, output_dir: Path, args) -> bool:
    print(f"\n  [{csv_path.name}]")

    raw_bytes = csv_path.read_bytes()
    if not raw_bytes:
        print("  SKIPPED — file is empty.")
        return False

    enc = detect_encoding(raw_bytes)
    try:
        sample = raw_bytes[:8192].decode(enc, errors="replace")
    except Exception:
        sample = raw_bytes[:8192].decode("utf-8", errors="replace")
        enc = "utf-8"
    delim = detect_delimiter(sample)

    # Parse CSV with multiple fallbacks
    df_raw = None
    for try_enc, try_delim in [(enc, delim), (enc, ","), ("utf-8", delim), ("latin-1", delim), ("latin-1", ",")]:
        try:
            df_raw = pd.read_csv(
                io.BytesIO(raw_bytes),
                encoding=try_enc, sep=try_delim,
                on_bad_lines="skip", low_memory=False,
            )
            enc, delim = try_enc, try_delim
            break
        except Exception:
            continue

    if df_raw is None or df_raw.empty:
        print("  SKIPPED — could not read or no usable data.")
        return False

    score = quality_score(df_raw, [
        c for c in df_raw.columns
        if df_raw[c].isna().sum() / max(len(df_raw), 1) * 100 >= args.col_threshold
    ])
    print(f"  Quality score (before) : {score}/100")
    print(f"  Rows: {len(df_raw):,}  |  Cols: {len(df_raw.columns)}")

    try:
        df, report = clean_dataframe(df_raw, args)
    except ValueError as e:
        print(f"  ERROR: {e}")
        return False
    except Exception as e:
        print(f"  ERROR during cleaning: {e}")
        return False

    if df.empty:
        print("  SKIPPED — no rows remain after cleaning. Try a fill strategy.")
        return False

    stamp    = datetime.now().strftime("%Y%m%d")
    safe     = sanitize_filename(csv_path.stem)
    xlsx_out = output_dir / f"{safe}_{stamp}.xlsx"
    log_out  = output_dir / f"{safe}_{stamp}_log.txt"

    try:
        write_excel(df, args.sheet, xlsx_out)
    except Exception as e:
        print(f"  ERROR writing Excel: {e}")
        return False

    print(f"  Rows: {report['rows_before']:,} → {report['rows_after']:,}  |  Cols: {report['cols_before']} → {report['cols_after']}")

    if report["dropped_cols"]:
        for col, pct in report["dropped_cols"]:
            lbl = "fully empty" if pct == 100 else f"{pct}% missing"
            print(f"  Column '{col}' dropped ({lbl})")

    if report["coerced_cols"]:
        print(f"  Type-coerced to numeric: {', '.join(report['coerced_cols'])}")

    if report["inf_cols"]:
        print(f"  Infinite values fixed in: {', '.join(report['inf_cols'])}")

    if report["imputed_cols"]:
        print(f"  Imputed {len(report['imputed_cols'])} column(s) via {args.strategy}")

    print(f"  Saved → {xlsx_out.name}")

    if not args.no_log:
        write_log(csv_path.name, report, score, enc, delim, log_out)
        print(f"  Log   → {log_out.name}")

    return True


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="DataClean Pro — CSV to Excel with intelligent cleaning",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("files", nargs="*", help="CSV file(s) to process")
    parser.add_argument("--input",          default=".",          help="Input folder (default: current dir)")
    parser.add_argument("--output",         default="output",     help="Output folder (default: ./output)")
    parser.add_argument("--strategy",       default="fill_median",
                        choices=["fill_median", "fill_mean", "fill_zero", "fill_fixed", "drop_row", "drop_empty"])
    parser.add_argument("--fixed-value",    default="0",          help="Value for --strategy fill_fixed")
    parser.add_argument("--col-threshold",  type=int, default=100)
    parser.add_argument("--keep-duplicates", action="store_true")
    parser.add_argument("--no-log",          action="store_true")
    parser.add_argument("--sheet",           default="Cleaned Data")
    args = parser.parse_args()

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.files:
        csv_files = [Path(f) for f in args.files if Path(f).suffix.lower() == ".csv" and Path(f).exists()]
        missing   = [f for f in args.files if not Path(f).exists()]
        for m in missing:
            print(f"  WARNING: file not found — {m}")
    else:
        input_dir = Path(args.input)
        if not input_dir.exists():
            print(f"ERROR: input folder not found: {input_dir}")
            sys.exit(1)
        csv_files = list(input_dir.glob("*.csv"))

    if not csv_files:
        print("No CSV files found.")
        sys.exit(1)

    print(f"\n{'='*55}")
    print(f"  DataClean Pro  v3.0")
    print(f"  Files    : {len(csv_files)}")
    print(f"  Strategy : {args.strategy}")
    print(f"  Threshold: drop cols with >= {args.col_threshold}% missing")
    print(f"  Output   : {output_dir.resolve()}")
    print(f"{'='*55}")

    ok = sum(process_file(f, output_dir, args) for f in csv_files)

    print(f"\n{'='*55}")
    print(f"  Done — {ok}/{len(csv_files)} file(s) processed successfully.")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
